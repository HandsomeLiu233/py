import xlwt
import openpyxl
from openpyxl import Workbook
wb=Workbook()

# 声明文件名称，方便用
txtname = 'f1.txt'
excelname = 't2e3st.xls'

# 读取txt文件
fopen = open(txtname, 'r')
lines = fopen.readlines()

#创建工作表，增加表头
ws=wb.create_sheet(title='武汉市2019年高考成绩')
ws['A1']='考号'
ws['B1']='姓名'
ws['C1']='性别'
ws['D1']='成绩'


for line in lines:
    line=line.strip().split(' ')
    ws.append(line)


wb.save(excelname)



